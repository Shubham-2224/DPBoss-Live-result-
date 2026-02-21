import requests
from bs4 import BeautifulSoup
import pandas as pd
import schedule
import time
import os
import re
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://dpboss.boston/"
FILE_NAME = r"C:\Users\SHUBHAM\OneDrive\Desktop\dpboss live result\live_data.xlsx"

def scrape_data():
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    try:
        response = requests.get(URL, headers=headers, timeout=20)
        soup = BeautifulSoup(response.text, 'html.parser')
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

    all_data = []
    
    # Track if we have reached the live result section
    start_collecting = False
    
    h4_tags = soup.find_all('h4')
    for h4 in h4_tags:
        name = h4.get_text(strip=True)
        
        # Marker to start collecting markets
        if "WORLD ME SABSE FAST" in name.upper():
            start_collecting = True
            continue
            
        if not start_collecting:
            continue
            
        # Stop marker (common sections after the live list)
        if any(stop_word in name.upper() for stop_word in ["FREE GAME", "JODI LIST", "CHART LIST", "GUESSING"]):
            break
            
        # Extract the h4 text and all its following siblings until the next h4
        # This ensures we get the results (often in <span>) and times (often in <p>)
        block_text = h4.get_text("|", strip=True)
        for sibling in h4.find_next_siblings():
            if sibling.name == 'h4':
                break
            # Also stop if we hit a stop marker in a sibling
            if sibling.name and any(stop in sibling.get_text().upper() for stop in ["FREE GAME", "JODI LIST", "CHART LIST", "GUESSING"]):
                break
            block_text += "|" + sibling.get_text("|", strip=True)
        
        # Format usually: Name | Result | Times | Jodi | Panel
        parts = block_text.split("|")
        result_str = ""
        times_str = ""
        
        # Look for result patterns
        for part in parts:
            part_strip = part.strip()
            # This regex looks for 3 digits, a hyphen, 1-2 digits, and optionally another hyphen and 3 digits
            # It matches '123-4', '123-45', '123-45-678', etc.
            if re.search(r'[\d\*]{3}-[\d\*]{1,2}', part_strip):
                result_str = part_strip
            elif re.search(r'\d{1,2}:\d{2}', part_strip):
                times_str = part_strip
        
        # Split result into Open, Main, Close
        open_panna = ""
        main_jodi = ""
        close_panna = ""
        
        if result_str:
            # We split by hyphens manually to be very explicit (e.g. "123-45-789" or "123-4")
            res_parts = [r.strip() for r in result_str.split("-")]
            
            # Put whatever parts we have into the respective columns
            if len(res_parts) >= 1: open_panna = res_parts[0]
            if len(res_parts) >= 2: main_jodi = res_parts[1]
            if len(res_parts) >= 3: close_panna = res_parts[2]
            
            # Clean parts of any non-numeric or non-asterisk text (like 'MORNING' or 'Close:')
            open_panna = re.sub(r'[^\d\*]', '', open_panna)
            main_jodi = re.sub(r'[^\d\*]', '', main_jodi)
            close_panna = re.sub(r'[^\d\*]', '', close_panna)
            
        open_time = ""
        close_time = ""
        if times_str:
            time_matches = re.findall(r'\d{1,2}:\d{2}\s?(?:AM|PM)', times_str, re.I)
            if len(time_matches) >= 1: open_time = time_matches[0]
            if len(time_matches) >= 2: close_time = time_matches[1]

        all_data.append({
            "Name": name,
            "Open": open_panna,
            "Main": main_jodi,
            "Close": close_panna,
            "Open Time": open_time,
            "Close Time": close_time
        })

    if not all_data:
        print("No live results found. The start marker 'WORLD ME SABSE FAST' might have changed.")
        return None

    return pd.DataFrame(all_data)

def apply_cell_style(cell, is_header=False):
    """Apply standard styling to a single cell."""
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    cell.alignment = align_center
    cell.border = thin_border
    
    if is_header:
        header_fill = PatternFill(patternType='solid', fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = header_fill
        cell.font = header_font

def update_excel():
    timestamp = time.strftime('%H:%M:%S')
    
    # Ensure the directory exists
    dir_name = os.path.dirname(FILE_NAME)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)
        print(f"📁 Created directory: {dir_name}")
        
    print(f"🔄 [{timestamp}] Fetching latest data...")
    
    new_df = scrape_data()
    if new_df is None or new_df.empty:
        print(f"⚠️ [{timestamp}] No data found on website. Skipping update.")
        return

    yellow_fill = PatternFill(patternType='solid', fgColor='FFFF00')
    no_fill = PatternFill(fill_type=None)

    total_changes = 0
    try:
        if os.path.exists(FILE_NAME):
            try:
                old_df = pd.read_excel(FILE_NAME, dtype=str).fillna("")
                wb = load_workbook(FILE_NAME)
                ws = wb.active
            except Exception as e:
                print(f"❌ Could not read old file: {e}. Creating fresh.")
                new_df.to_excel(FILE_NAME, index=False)
                return

            old_map = {str(row['Name']).strip(): row for _, row in old_df.iterrows()}
            
            # Update columns just in case
            for col_num, col_name in enumerate(new_df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = col_name
                apply_cell_style(cell, is_header=True)
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            for row_idx, new_row in new_df.iterrows():
                name = str(new_row['Name']).strip()
                excel_row = row_idx + 2
                
                for col_idx, col_name in enumerate(new_df.columns):
                    val = new_row[col_name]
                    excel_col = col_idx + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    
                    old_val = ""
                    if name in old_map:
                        old_val = str(old_map[name].get(col_name, "")).strip()
                    
                    new_val_str = str(val).strip()
                    if new_val_str == "None": new_val_str = ""
                    
                    cell.value = new_val_str
                    apply_cell_style(cell)
                    
                    if name in old_map and new_val_str != old_val and new_val_str not in ["", "***", "**"]:
                        # Special log for high-interest markets
                        if "PARVATI" in name.upper() or "MAHADEVI" in name.upper():
                            print(f"✨ UPDATE: {name} - {col_name}: '{old_val}' -> '{new_val_str}'")
                        cell.fill = yellow_fill
                        total_changes += 1
                    else:
                        cell.fill = no_fill

            wb.save(FILE_NAME)
            print(f"✅ [{timestamp}] Update complete! {total_changes} changes highlighted.")
            
        else:
            new_df.to_excel(FILE_NAME, index=False)
            wb = load_workbook(FILE_NAME)
            ws = wb.active
            for col_num, col_name in enumerate(new_df.columns, 1):
                apply_cell_style(ws.cell(row=1, column=col_num), is_header=True)
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            for row_idx in range(len(new_df)):
                for col_idx in range(len(new_df.columns)):
                    apply_cell_style(ws.cell(row=row_idx+2, column=col_idx+1))
            wb.save(FILE_NAME)
            print(f"✅ [{timestamp}] Created new Excel: {FILE_NAME}")

    except PermissionError:
        print(f"❌ ERROR: Could not save '{FILE_NAME}'. Please CLOSE the Excel file!")
    except Exception as e:
        print(f"❌ Unexpected error during update: {e}")

if __name__ == "__main__":
    update_excel()
    schedule.every(1).minutes.do(update_excel)
    print("Auto-update started (1 min intervals). Press Ctrl+C to exit.")
    while True:
        schedule.run_pending()
        time.sleep(1)