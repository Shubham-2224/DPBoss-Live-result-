from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('live_data.xlsx')
ws = wb.active

highlighted_cells = []
for row in ws.iter_rows():
    for cell in row:
        if cell.fill and cell.fill.fill_type is not None:
            print(f"Cell {cell.coordinate}: fill_type={cell.fill.fill_type}, fgColor={cell.fill.fgColor}")
            if cell.fill.fgColor and cell.fill.fgColor.rgb in ['00FFFF00', 'FFFF00', 'FFFFFF00']:
                highlighted_cells.append((cell.coordinate, cell.fill.fgColor.rgb))

print(f"Found {len(highlighted_cells)} highlighted cells.")
for coord, color in highlighted_cells:
    print(f"Cell {coord}: {color}")
